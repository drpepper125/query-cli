import json
import boto3
import logging
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from rich import print
from aws_helper import create_mgmt_session, create_read_only_session

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


def read_account_lists(file_path):
    """Read account lists from JSON file"""
    with open(file_path, 'r') as file:
        account_lists = json.load(file)
    return account_lists


def write_to_workbook(workbook, data, env):
    """
    Write certificate data to a worksheet in the workbook.
    
    Args:
        workbook: openpyxl.Workbook object
        data: List of dictionaries containing certificate data
        env: Environment name to use as worksheet name (e.g., "dev-finops")
    """
    if not data:
        logger.info(f"No data to write for {env}, skipping worksheet creation")
        return
    
    # Create new worksheet with the environment name
    worksheet = workbook.create_sheet(title=env)
    
    # Extract column headers from first certificate dict keys
    headers = list(data[0].keys())
    
    # Write headers to row 1
    for col, header in enumerate(headers, 1):
        worksheet.cell(row=1, column=col, value=header)
    
    # Write certificate data rows starting from row 2
    for row_idx, cert_data in enumerate(data, 2):
        for col_idx, header in enumerate(headers, 1):
            value = cert_data.get(header, '')
            # Convert datetime objects to string
            if hasattr(value, 'isoformat'):
                value = value.isoformat()
            worksheet.cell(row=row_idx, column=col_idx, value=str(value))
    
    # Auto-adjust column widths for readability
    for column in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[column_letter].width = adjusted_width
    
    logger.info(f"Wrote {len(data)} certificates to worksheet '{env}'")


def get_certificates_for_account(mgmt_session, account_id, account_name, role_name, region='us-east-1'):
    """
    Get certificates for a specific account.
    
    Args:
        mgmt_session: Management boto3 session
        account_id: AWS account ID
        account_name: Account name/alias
        role_name: IAM role name to assume
        region: AWS region (default: us-east-1)
        
    Returns:
        Tuple of (account_name, certificates_list)
    """
    try:
        # Create read-only session
        read_only_session = create_read_only_session(mgmt_session, account_id, role_name)
        
        # Create ACM client
        acm_client = read_only_session.client('acm', region_name=region)
        
        # Use paginator to get all certificates
        paginator = acm_client.get_paginator('list_certificates')
        certificates = []
        
        for page in paginator.paginate():
            for certificate in page['CertificateSummaryList']:
                certificates.append(certificate)
        
        logger.info(f"Found {len(certificates)} certificates in {account_name} ({account_id})")
        return (account_name, certificates)
        
    except Exception as e:
        logger.error(f"Error getting certificates for {account_name} ({account_id}): {e}")
        return (account_name, [])


def gather_product_certificates(account_lists_path, target_env, role_name='read-only-role'):
    """
    Gather certificates for all products and create Excel workbooks.
    
    Args:
        account_lists_path: Path to account_lists.json file
        target_env: Target environment ('pro' or 'dev'/'pre'/'poc')
        role_name: IAM role name to assume (default: 'read-only-role')
    """
    # Read account lists from JSON
    account_lists = read_account_lists(account_lists_path)
    
    # Create root STS client
    root_sts_client = boto3.client('sts')
    
    # Process each product
    for product, accounts in account_lists.items():
        print(f"\n[bold green]Processing product: {product}[/bold green]")
        
        # Filter accounts based on target_env
        filtered_accounts = {}
        for account_id, account_name in accounts.items():
            if target_env == 'pro':
                # Skip dev, pre, poc accounts
                if not any(env in account_name.lower() for env in ['dev', 'pre', 'poc']):
                    filtered_accounts[account_id] = account_name
            elif target_env in ['dev', 'pre', 'poc']:
                # Skip pro accounts
                if 'pro' not in account_name.lower():
                    filtered_accounts[account_id] = account_name
        
        if not filtered_accounts:
            print(f"  [yellow]No accounts to process for environment '{target_env}'[/yellow]")
            continue
        
        print(f"  Processing {len(filtered_accounts)} accounts for environment '{target_env}'")
        
        # Create management session
        try:
            mgmt_session = create_mgmt_session(root_sts_client, target_env)
        except Exception as e:
            logger.error(f"Failed to create management session for {product}: {e}")
            continue
        
        # Create workbook for this product
        workbook = Workbook()
        # Remove default sheet
        if 'Sheet' in workbook.sheetnames:
            workbook.remove(workbook['Sheet'])
        
        # Use ThreadPoolExecutor for concurrent processing
        with ThreadPoolExecutor(max_workers=5) as executor:
            # Submit tasks for each account
            future_to_account = {
                executor.submit(
                    get_certificates_for_account,
                    mgmt_session,
                    account_id,
                    account_name,
                    role_name
                ): account_name
                for account_id, account_name in filtered_accounts.items()
            }
            
            # Collect results as they complete
            for future in as_completed(future_to_account):
                account_name = future_to_account[future]
                try:
                    account_name_result, certificates = future.result()
                    
                    if certificates:
                        print(f"  [blue]✓[/blue] {account_name_result}: {len(certificates)} certificates")
                        write_to_workbook(workbook, certificates, account_name_result)
                    else:
                        print(f"  [dim]- {account_name_result}: No certificates[/dim]")
                        
                except Exception as e:
                    logger.error(f"Error processing {account_name}: {e}")
                    print(f"  [red]✗[/red] {account_name}: Error occurred")
        
        # Save workbook only if it has worksheets
        if workbook.sheetnames:
            filename = f"{product}_certificates.xlsx"
            workbook.save(filename)
            print(f"[green]✓ Created workbook: {filename}[/green]")
        else:
            print(f"[yellow]No certificates found for {product}, skipping workbook creation[/yellow]")


def main():
    """Main function to run certificate gathering"""
    import sys
    
    # Default to 'pro' environment if not specified
    target_env = sys.argv[1] if len(sys.argv) > 1 else 'pro'
    
    if target_env not in ['pro', 'dev', 'pre', 'poc']:
        print(f"[red]Invalid environment: {target_env}[/red]")
        print("Valid environments: pro, dev, pre, poc")
        sys.exit(1)
    
    print(f"[bold]Gathering certificates for environment: {target_env}[/bold]\n")
    gather_product_certificates('account_lists.json', target_env)
    print("\n[bold green]Certificate gathering complete![/bold green]")


if __name__ == "__main__":
    main()
